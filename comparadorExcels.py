from tkinter import Tk, Label, filedialog, Button
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from difflib import SequenceMatcher


def find_header_row(file, keywords):
    # Intentar leer las primeras 10 filas del archivo con pandas sin definir el encabezado
    try:
        df = pd.read_excel(file, engine="openpyxl", nrows=10, header=None)

        # Verificar si la primera fila contiene alguna de las palabras clave
        for i, cell in enumerate(df.iloc[0]):
            if any(keyword.lower() in str(cell).lower() for keyword in keywords):
                return 1  # El encabezado está en la primera fila

        # Si no se encuentra en la primera fila, buscar el encabezado manualmente
        for i, row in df.iterrows():
            if any(any(keyword.lower() in str(cell).lower() for keyword in keywords) for cell in row):
                return i + 1  # Retorna la fila del encabezado (1-indexed)
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
        return -1  # Si no se puede leer el archivo o no se encuentra el encabezado

    # Si no se encuentra el encabezado, devolver -1
    return -1


def compare_excels(file1, file2, output_file):
    # Lista de palabras clave para buscar en los encabezados
    keywords = ["name", "nombre", "pass", "영문명", "영문이름", "영문성함", "여권번호"]

    # Detectar la fila de encabezados en ambos archivos usando las palabras clave
    header_row1 = find_header_row(file1, keywords)
    header_row2 = find_header_row(file2, keywords)

    if header_row1 == -1 or header_row2 == -1:
        print("No se encontró una fila de encabezado válida en uno de los archivos.")
        return

    # Leer ambos archivos Excel desde la fila detectada
    df1 = pd.read_excel(file1, header=header_row1 - 1, engine="openpyxl")
    df2 = pd.read_excel(file2, header=header_row2 - 1, engine="openpyxl")

    # Renombrar columnas no válidas
    df1.columns = [f"Col_{i}" if 'Unnamed' in col else col for i, col in enumerate(df1.columns)]
    df2.columns = [f"Col_{i}" if 'Unnamed' in col else col for i, col in enumerate(df2.columns)]

    # Identificar columnas de interés
    cols_of_interest1 = [col for col in df1.columns if any(
        keyword in col.lower() for keyword in keywords) and "kor" not in col.lower()]
    cols_of_interest2 = [col for col in df2.columns if any(
        keyword in col.lower() for keyword in keywords) and "kor" not in col.lower()]

    # Filtrar ambas tablas y verificar que las columnas de interés existen
    if len(cols_of_interest1) == 0 or len(cols_of_interest2) == 0:
        print("No se encontraron las columnas de interés en uno de los archivos.")
        return

    df1_filtered = df1[cols_of_interest1]
    df2_filtered = df2[cols_of_interest2]

    # Verificar que las columnas filtradas tengan el tamaño esperado
    if df1_filtered.shape[1] != 2 or df2_filtered.shape[1] != 2:
        print("El número de columnas en los archivos filtrados no es el esperado.")
        return

    # Normalizar nombres de columnas
    df1_filtered.columns = df2_filtered.columns = ["Nombre/Apellido", "Pasaporte"]

    # Crear diccionarios para comparación
    dict1 = {str(row["Nombre/Apellido"]).strip().lower(): str(row["Pasaporte"]).strip() for _, row in df1_filtered.iterrows()}
    dict2 = {str(row["Nombre/Apellido"]).strip().lower(): str(row["Pasaporte"]).strip() for _, row in df2_filtered.iterrows()}

    # Función para calcular similitud entre cadenas
    def nombres_similares(nombre1, nombre2, umbral=0.85):
        return SequenceMatcher(None, nombre1, nombre2).ratio() >= umbral

    # Inicializar las filas de resultado respetando el orden del archivo 1
    result_rows = []
    for name1 in dict1:
        passport1 = dict1.get(name1, "Vacío")

        # Verificar si el nombre o pasaporte está vacío (NaN)
        if pd.isna(name1) or pd.isna(passport1):
            continue  # Si está vacío, saltar esta fila

        match_found = False

        for name2 in dict2:
            # Verificar si los nombres son similares
            if nombres_similares(name1, name2):
                passport2 = dict2[name2]

                # Si los pasaportes son diferentes, marcar como "MODIFICADO"
                if passport1 != passport2:
                    estado = "MODIFICADO"
                    nombre_diff = f"{name1.title()} -> {name2.title()}"
                    result_rows.append([nombre_diff, f"{passport1} -> {passport2}", estado])
                    match_found = True
                    break
                else:  # Si el nombre y el pasaporte coinciden
                    estado = "MANTENIDO"
                    nombre_diff = name1.title()
                    result_rows.append([nombre_diff, passport1, estado])
                    match_found = True
                    break

        if not match_found:
            result_rows.append([f"{name1.title()} -> Vacío", f"{passport1} -> Vacío", "CANCELADO"])

    # Añadir las filas exclusivas del archivo 2 (fuera del orden del archivo 1)
    for name2 in dict2:
        if not any(nombres_similares(name2, name1) for name1 in dict1):
            passport2 = dict2[name2]

            # Verificar si el nombre o pasaporte está vacío (NaN)
            if pd.isna(name2) or pd.isna(passport2):
                continue  # Si está vacío, saltar esta fila

            result_rows.append([f"Vacío -> {name2.title()}", f"Vacío -> {passport2}", "NUEVO"])

    # Crear DataFrame de diferencias con columna Estado
    diff_df = pd.DataFrame(result_rows, columns=["Nombre/Apellido", "Pasaporte", "Estado"])

    # Guardar diferencias en un nuevo archivo
    diff_df.to_excel(output_file, index=False)

    # Resaltar diferencias en el archivo generado
    wb = load_workbook(output_file)
    ws = wb.active
    mod_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cancel_fill = PatternFill(start_color="FF5050", end_color="FF5050", fill_type="solid")
    new_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Resaltar diferencias en las celdas donde hay cambios
    for row in range(2, ws.max_row + 1):
        estado_cell = ws.cell(row=row, column=3)  # Columna de Estado
        nombre_cell = ws.cell(row=row, column=1)  # Columna de Nombre/Apellido
        pasaporte_cell = ws.cell(row=row, column=2)  # Columna de Pasaporte

        if estado_cell.value == "MODIFICADO":
            estado_cell.fill = mod_fill
            nombre_cell.fill = mod_fill
            pasaporte_cell.fill = mod_fill
        elif estado_cell.value == "NUEVO":
            estado_cell.fill = new_fill
            nombre_cell.fill = new_fill
            pasaporte_cell.fill = new_fill
        elif estado_cell.value == "CANCELADO":
            estado_cell.fill = cancel_fill
            nombre_cell.fill = cancel_fill
            pasaporte_cell.fill = cancel_fill

    # Guardar el archivo final
    wb.save(output_file)


def browse_file(label):
    filename = filedialog.askopenfilename(filetypes=[["Excel files", "*.xlsx"]])
    if filename:
        label.config(text=filename)
        return filename
    return None


def run_comparator():
    file1 = label_file1.cget("text")
    file2 = label_file2.cget("text")

    if not file1 or not file2:
        result_label.config(text="Por favor, selecciona ambos archivos.", fg="red")
        return

    output_file = f"roomingConDiferencias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        compare_excels(file1, file2, output_file)
        result_label.config(text=f"Comparación completada. Archivo guardado como: {output_file}", fg="green")
    except ValueError as e:
        result_label.config(text=str(e), fg="red")


# Crear la interfaz gráfica
root = Tk()
root.title("Comparador de Excel de Roomings")

Label(root, text="Selecciona el Excel 1").pack()
label_file1 = Label(root, text="", bg="lightgray", width=50)
label_file1.pack(pady=5)
Button(root, text="Seleccionar Excel 1", command=lambda: browse_file(label_file1)).pack()

Label(root, text="Selecciona el Excel 2").pack()
label_file2 = Label(root, text="", bg="lightgray", width=50)
label_file2.pack(pady=5)
Button(root, text="Seleccionar Excel 2", command=lambda: browse_file(label_file2)).pack()

Button(root, text="Comparar Excels", command=run_comparator).pack(pady=20)

result_label = Label(root, text="")
result_label.pack()

root.mainloop()
